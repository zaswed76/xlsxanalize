

import os
from openpyxl import load_workbook
# xls_dir = './data'
# xlsfile = os.path.join(xls_dir, '01.10.16 - 02.10.2016 (24).xlsx')
# print(xlsfile)
#
#
# wb2 = load_workbook(xlsfile)
# sheet = wb2[wb2.get_sheet_names()[0]]
# name = sheet['C1'].value
# date = sheet['C3'].value
# print(name)
# print(date)


def list_paths(direct):
    return [os.path.join(direct, x) for x in os.listdir(direct)]

class Xls:
    def __init__(self):
        pass

    def read(self, lst):
        book = []
        for i in lst:
            d = {}

            # d.clear()
            wb = load_workbook(i)
            sheet = wb.active
            print(sheet, 111)
            d['name'] = sheet['C1'].value
            print(d['name'])
            d['date'] = sheet['C3'].value
            book.append(d)
        return book




if __name__ == '__main__':
    xls_dir = './data'
    xls = Xls()
    print(list_paths(xls_dir))
    print(xls.read(list_paths(xls_dir)))